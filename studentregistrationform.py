##Integrating Student Registration Details taken in Tkinter GUI using python and Updating the Details on Excel Sheet

from openpyxl import *
from tkinter import *
  


wb = load_workbook(r"C:/Users/Arnab Dutta/Desktop/Book1.xlsx") 

sheet = wb.active 
  
  
def excel(): 
      
 
    sheet.column_dimensions['A'].width = 100
    sheet.column_dimensions['B'].width = 100
    sheet.column_dimensions['C'].width = 100
    sheet.column_dimensions['D'].width = 100
    sheet.column_dimensions['E'].width = 100
    sheet.column_dimensions['F'].width = 100
    sheet.column_dimensions['G'].width = 100
    sheet.column_dimensions['H'].width = 100
    sheet.column_dimensions['I'].width = 100
    sheet.column_dimensions['J'].width = 100
    sheet.column_dimensions['K'].width = 100
    sheet.column_dimensions['L'].width = 100
    sheet.column_dimensions['M'].width = 100
    sheet.column_dimensions['N'].width = 100
    sheet.column_dimensions['O'].width = 100
    sheet.column_dimensions['P'].width = 100
    sheet.column_dimensions['Q'].width = 100
    sheet.column_dimensions['R'].width = 100
    sheet.column_dimensions['S'].width = 100
    sheet.column_dimensions['T'].width = 100 
    sheet.column_dimensions['U'].width = 100
    sheet.column_dimensions['V'].width = 100
    sheet.column_dimensions['W'].width = 100
    sheet.column_dimensions['X'].width = 100
  
    
    sheet.cell(row=1, column=1).value = "student_name"
    sheet.cell(row=1, column=2).value = "student_age"
    sheet.cell(row=1, column=3).value = "student_dob"
    sheet.cell(row=1, column=4).value = "student_primary_contact_number"
    sheet.cell(row=1, column=5).value = "student_secondary_contact_number"
    sheet.cell(row=1, column=6).value = "student_permanent_address"
    sheet.cell(row=1, column=7).value = "student_temporary_address"
    sheet.cell(row=1, column=8).value = "student_father_name"
    sheet.cell(row=1, column=9).value = "student_mother_name"
    sheet.cell(row=1, column=10).value = "student_email_id"
    sheet.cell(row=1, column=11).value = "student_gender"
    sheet.cell(row=1, column=12).value = "student_religion"
    sheet.cell(row=1, column=13).value = "student_nationality"
    sheet.cell(row=1, column=14).value = "student_languages_known"
    sheet.cell(row=1, column=15).value = "student_10th_marks"
    sheet.cell(row=1, column=16).value = "student_12th_marks"
    sheet.cell(row=1, column=17).value = "student_entrance_rank_jee"
    sheet.cell(row=1, column=18).value = "student_any_medical_history"
    sheet.cell(row=1, column=19).value = "student_medical_history_description"
    sheet.cell(row=1, column=20).value = "student_hobbies_and_interests"
    sheet.cell(row=1, column=21).value = "student_first_choice_department"
    sheet.cell(row=1, column=22).value = "student_second_choice_department"
    sheet.cell(row=1, column=23).value = "student_third_choice_departmentt"
    sheet.cell(row=1, column=24).value = "student_any_other_details"
  
def focus1(event): 
    student_name_field.focus_set() 
 
def focus2(event):   
    student_age_field.focus_set() 

def focus3(event):   
    student_dob_field.focus_set() 
 
def focus4(event):  
    student_primary_contact_number_field.focus_set() 
  
def focus5(event):
    student_secondary_contact_number_field.focus_set() 
  
def focus6(event): 
    student_permanent_address_field.focus_set() 

def focus7(event): 
    student_temporary_address_field.focus_set() 

def focus8(event):    
    student_father_name_field.focus_set() 
 
def focus9(event):   
    student_mother_name_field.focus_set() 
  
def focus10(event):     
    student_email_id_field.focus_set() 

def focus11(event):    
    student_gender_field.focus_set() 

def focus12(event):   
    student_religion_field.focus_set() 
 
def focus13(event):   
    student_nationality_field.focus_set() 
  
def focus14(event):   
    student_languages_known_field.focus_set() 
  
def focus15(event):    
    student_10th_marks_field.focus_set() 
   
def focus16(event):   
    student_12th_marks_field.focus_set()  
  
def focus17(event):    
    student_entrance_rank_jee_field.focus_set() 

def focus18(event): 
    student_any_medical_history_field.focus_set() 

def focus19(event):   
    student_medical_history_description_field.focus_set() 
  
def focus20(event): 
    student_hobbies_and_interests_field.focus_set() 

def focus21(event):  
    student_first_choice_department_field.focus_set() 
   
def focus22(event):  
    student_second_choice_department_field.focus_set() 

def focus23(event): 
    student_third_choice_departmentt_field.focus_set()
 
def focus24(event): 
    student_any_other_details_field.focus_set()

def clear(): 
      
    student_name_field.delete(0, END) 
    student_age_field.delete(0, END) 
    student_dob_field.delete(0, END) 
    student_primary_contact_number_field.delete(0, END) 
    student_secondary_contact_number_field.delete(0, END) 
    student_permanent_address_field.delete(0, END) 
    student_temporary_address_field.delete(0, END)
    student_father_name_field.delete(0, END) 
    student_mother_name_field.delete(0, END) 
    student_email_id_field.delete(0, END) 
    student_gender_field.delete(0, END) 
    student_religion_field.delete(0, END) 
    student_nationality_field.delete(0, END) 
    student_languages_known_field.delete(0, END)
    student_10th_marks_field.delete(0, END) 
    student_12th_marks_field.delete(0, END) 
    student_entrance_rank_jee_field.delete(0, END) 
    student_any_medical_history_field.delete(0, END) 
    student_medical_history_description_field.delete(0, END) 
    student_hobbies_and_interests_field.delete(0, END) 
    student_first_choice_department_field.delete(0, END)
    student_second_choice_department_field.delete(0, END) 
    student_third_choice_departmentt_field.delete(0, END) 
    student_any_other_details_field.delete(0, END)
   
def insert(): 
      
   
    if (student_name_field.get() == "" and
        student_age_field.get() == "" and
        student_dob_field.get() == "" and
        student_primary_contact_number_field.get() == "" and
        student_secondary_contact_number_field.get() == "" and
        student_permanent_address_field.get() == "" and
        student_temporary_address_field.get() == "" and
        student_father_name_field.get() == "" and
        student_mother_name_field.get() == "" and
        student_email_id_field.get() == "" and
        student_gender_field.get() == "" and
        student_religion_field.get() == "" and
        student_nationality_field.get() == "" and
        student_languages_known_field.get() == "" and
        student_10th_marks_field.get() == "" and
        student_12th_marks_field.get() == "" and
        student_entrance_rank_jee_field.get() == "" and
        student_any_medical_history_field.get() == "" and
        student_medical_history_description_field.get() == "" and
        student_hobbies_and_interests_field.get() == "" and
        student_first_choice_department_field.get() == "" and
        student_second_choice_department_field.get() == "" and
        student_third_choice_departmentt_field.get() == "" and
        student_any_other_details_field.get() == ""): 
              
        print("empty input") 
  
    else: 
  
        current_row = sheet.max_row 
        current_column = sheet.max_column 
  
      
        sheet.cell(row=current_row + 1, column=1).value = student_name_field.get() 
        sheet.cell(row=current_row + 1, column=2).value = student_age_field.get() 
        sheet.cell(row=current_row + 1, column=3).value = student_dob_field.get()
        sheet.cell(row=current_row + 1, column=4).value = student_primary_contact_number_field.get() 
        sheet.cell(row=current_row + 1, column=5).value = student_secondary_contact_number_field.get() 
        sheet.cell(row=current_row + 1, column=6).value = student_permanent_address_field.get() 
        sheet.cell(row=current_row + 1, column=7).value = student_temporary_address_field.get() 
        sheet.cell(row=current_row + 1, column=8).value = student_father_name_field.get() 
        sheet.cell(row=current_row + 1, column=9).value = student_mother_name_field.get() 
        sheet.cell(row=current_row + 1, column=10).value = student_email_id_field.get() 
        sheet.cell(row=current_row + 1, column=11).value = student_gender_field.get()
        sheet.cell(row=current_row + 1, column=12).value = student_religion_field.get()
        sheet.cell(row=current_row + 1, column=13).value = student_nationality_field.get()
        sheet.cell(row=current_row + 1, column=14).value = student_languages_known_field.get()
        sheet.cell(row=current_row + 1, column=15).value = student_10th_marks_field.get() 
        sheet.cell(row=current_row + 1, column=16).value = student_12th_marks_field.get() 
        sheet.cell(row=current_row + 1, column=17).value = student_entrance_rank_jee_field.get() 
        sheet.cell(row=current_row + 1, column=18).value = student_any_medical_history_field.get() 
        sheet.cell(row=current_row + 1, column=19).value = student_medical_history_description_field.get() 
        sheet.cell(row=current_row + 1, column=20).value = student_hobbies_and_interests_field.get() 
        sheet.cell(row=current_row + 1, column=21).value = student_first_choice_department_field.get()
        sheet.cell(row=current_row + 1, column=22).value = student_second_choice_department_field.get() 
        sheet.cell(row=current_row + 1, column=23).value = student_third_choice_departmentt_field.get() 
        sheet.cell(row=current_row + 1, column=24).value = student_any_other_details_field.get()
  
  
        wb.save(r"C:/Users/Arnab Dutta/Desktop/Book1.xlsx") 
  
        student_name_field.focus_set() 
  
        clear() 
        print(student_any_other_details_field.get())
  
# Driver code 
if __name__ == "__main__": 

    root = Tk() 
  
    root.configure(background='azure') 
 
    root.title("Student Registration Form") 
  
    root.geometry("1105x295") 
  
    excel() 
  
    heading = Label(root, text="       Student Registration Form", bg="azure") 

    Name = Label(root, text="Name", bg="linen") 

    Age = Label(root, text="Age", bg="linen") 
  
    Date_Of_Birth = Label(root, text="Date Of Birth", bg="linen") 
  
    Primary_Contact_Number = Label(root, text="Primary Contact Number", bg="linen") 
  
    Secondary_Contact_Number = Label(root, text="Secondary Contact Number", bg="linen") 
  
    Permanent_Address = Label(root, text="Student Permanent Address", bg="linen") 

    Temporary_Address = Label(root, text="Student Temporary Address", bg="linen") 
     
    Father_Name = Label(root, text="Father's Name", bg="linen") 
     
    Mother_Name = Label(root, text="Mother's Name", bg="linen") 
  
    Email_Id = Label(root, text="Email ID", bg="linen") 
  
    Gender = Label(root, text="Gender", bg="linen") 
    
    Religion = Label(root, text="Religion", bg="linen") 
  
    Nationality = Label(root, text="Nationality", bg="linen") 
     
    Lanuages_Known = Label(root, text="Languages Known", bg="linen") 
  
    Marks_10th = Label(root, text="10th Marks", bg="linen") 
    
    Marks_12th = Label(root, text="12th Marks", bg="linen") 
  
    JEE_Entrance_Rank = Label(root, text="JEE Rank", bg="linen") 
     
    Any_Medical_History = Label(root, text="Any Medical History?", bg="linen") 
  
    Medical_History_Description = Label(root, text="Medical History Description", bg="linen") 
  
    Hobbies_and_Interests = Label(root, text="Hobbies and Interests", bg="linen") 

    First_Choice_Department = Label(root, text="First Choice Department", bg="linen") 
   
    Second_Choice_Department = Label(root, text="Second Choice Department", bg="linen") 
  
    Third_Choice_Department = Label(root, text="Third Choice Department", bg="linen") 
 
    Any_Other_Details = Label(root, text="Any Other Details", bg="linen")
    

    heading.grid(row=0, column=2) 
    Name.grid(row=1, column=0) 
    Age.grid(row=2, column=0) 
    Date_Of_Birth.grid(row=3, column=0) 
    Primary_Contact_Number.grid(row=4, column=0) 
    Secondary_Contact_Number.grid(row=5, column=0) 
    Permanent_Address.grid(row=6, column=0) 
    Temporary_Address.grid(row=7, column=0) 
    Father_Name.grid(row=8, column=0) 
    Mother_Name.grid(row=9, column=0) 
    Email_Id.grid(row=10, column=0) 
    Gender.grid(row=11, column=0) 
    Religion.grid(row=12, column=0) 
    Nationality.grid(row=1, column=4) 
    Lanuages_Known.grid(row=2, column=4) 
    Marks_10th.grid(row=3, column=4) 
    Marks_12th.grid(row=4, column=4) 
    JEE_Entrance_Rank.grid(row=5, column=4) 
    Any_Medical_History.grid(row=6, column=4) 
    Medical_History_Description.grid(row=7, column=4) 
    Hobbies_and_Interests.grid(row=8, column=4) 
    First_Choice_Department.grid(row=9, column=4) 
    Second_Choice_Department.grid(row=10, column=4) 
    Third_Choice_Department.grid(row=11, column=4) 
    Any_Other_Details.grid(row=12, column=4)  
  

    student_name_field = Entry(root)
    student_age_field = Entry(root)
    student_dob_field = Entry(root)
    student_primary_contact_number_field = Entry(root)
    student_secondary_contact_number_field = Entry(root)
    student_permanent_address_field = Entry(root)
    student_temporary_address_field = Entry(root)
    student_father_name_field = Entry(root)
    student_mother_name_field = Entry(root)
    student_email_id_field = Entry(root)
    student_gender_field = Entry(root)
    student_religion_field = Entry(root) 
    student_nationality_field = Entry(root)
    student_languages_known_field = Entry(root)
    student_10th_marks_field = Entry(root)
    student_12th_marks_field = Entry(root)
    student_entrance_rank_jee_field = Entry(root)
    student_any_medical_history_field = Entry(root)
    student_medical_history_description_field = Entry(root)
    student_hobbies_and_interests_field = Entry(root)
    student_first_choice_department_field = Entry(root)
    student_second_choice_department_field = Entry(root)
    student_third_choice_departmentt_field = Entry(root)
    student_any_other_details_field = Entry(root)
   
  
 
    student_name_field.bind("<Return>", focus1) 
 
    student_age_field.bind("<Return>", focus2) 
 
    student_dob_field.bind("<Return>", focus3) 
  
    student_primary_contact_number_field.bind("<Return>", focus4) 

    student_secondary_contact_number_field.bind("<Return>", focus5) 
  
    student_permanent_address_field.bind("<Return>", focus6)

    student_temporary_address_field.bind("<Return>", focus7) 

    student_father_name_field.bind("<Return>", focus8) 
   
    student_mother_name_field.bind("<Return>", focus9) 
  
    student_email_id_field.bind("<Return>", focus10) 
 
    student_gender_field.bind("<Return>", focus11) 
  
    student_religion_field.bind("<Return>", focus12)
 
    student_nationality_field.bind("<Return>", focus13) 
  
    student_languages_known_field.bind("<Return>", focus14) 
  
    student_10th_marks_field.bind("<Return>", focus15) 
   
    student_12th_marks_field.bind("<Return>", focus16) 
  
    student_entrance_rank_jee_field.bind("<Return>", focus17) 
   
    student_any_medical_history_field.bind("<Return>", focus18) 
 
    student_medical_history_description_field.bind("<Return>", focus19) 
  
    student_hobbies_and_interests_field.bind("<Return>", focus20) 
  
    student_first_choice_department_field.bind("<Return>", focus21)

    student_second_choice_department_field.bind("<Return>", focus22) 
  
    student_third_choice_departmentt_field.bind("<Return>", focus23) 

    student_any_other_details_field.bind("<Return>", focus24)
  
   
    student_name_field.grid(row=1, column=1, ipadx="100") 
    student_age_field.grid(row=2, column=1, ipadx="100") 
    student_dob_field.grid(row=3, column=1, ipadx="100") 
    student_primary_contact_number_field.grid(row=4, column=1, ipadx="100") 
    student_secondary_contact_number_field.grid(row=5, column=1, ipadx="100") 
    student_permanent_address_field.grid(row=6, column=1, ipadx="100") 
    student_temporary_address_field.grid(row=7, column=1, ipadx="100") 
    student_father_name_field.grid(row=8, column=1, ipadx="100") 
    student_mother_name_field.grid(row=9, column=1, ipadx="100") 
    student_email_id_field.grid(row=10, column=1, ipadx="100") 
    student_gender_field.grid(row=11, column=1, ipadx="100") 
    student_religion_field.grid(row=12, column=1, ipadx="100") 
    student_nationality_field.grid(row=1, column=5, ipadx="100") 
    student_languages_known_field.grid(row=2, column=5, ipadx="100")
    student_10th_marks_field.grid(row=3, column=5, ipadx="100") 
    student_12th_marks_field.grid(row=4, column=5, ipadx="100") 
    student_entrance_rank_jee_field.grid(row=5, column=5, ipadx="100") 
    student_any_medical_history_field.grid(row=6, column=5, ipadx="100") 
    student_medical_history_description_field.grid(row=7, column=5, ipadx="100") 
    student_hobbies_and_interests_field.grid(row=8, column=5, ipadx="100") 
    student_first_choice_department_field.grid(row=9, column=5, ipadx="100")
    student_second_choice_department_field.grid(row=10, column=5, ipadx="100") 
    student_third_choice_departmentt_field.grid(row=11, column=5, ipadx="100") 
    student_any_other_details_field.grid(row=12, column=5, ipadx="100") 
    
  
    
    excel() 
 
    submit = Button(root, text="Submit", fg="Black", 
                            bg="Red", command=insert) 
    submit.grid(row=25, column=2) 
  
  
    root.mainloop() 

